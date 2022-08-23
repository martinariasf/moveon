print("Python is running with openpyxl")
import openpyxl
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from tkinter import messagebox


# create the root window
messagebox.showinfo("Importing files", "First select the ASM file, after the DDM file and finally the MBA file")

def ASM_file():
    filetypes = (
        ('Excel files', '*.xlsx;*.xlsm'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='ASM file',
        initialdir='\\rsns01\maarias$\Desktop\test',
        filetypes=filetypes)

    showinfo(
        title='Selected ASM File: '+filename[-22:],
        message='Click Ok, wait and select the DDM file'
    )
    return filename

ASM_file_dir=ASM_file()
ASM=openpyxl.load_workbook(ASM_file_dir, data_only=True)

def DDM_file():
    filetypes = (
        ('Excel files', '*.xlsx;*.xlsm'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='DDM file',
        initialdir="\\rsns01\maarias$\Desktop\test",
        filetypes=filetypes)

    showinfo(
        title='Selected DDM File: '+filename[-22:],
        message='Click Ok, wait and select the MBA file'
    )
    return filename

DDM_file_dir=DDM_file()
DDM=openpyxl.load_workbook(DDM_file_dir, data_only=True)

def MBA_file():
    filetypes = (
        ('Excel files', '*.xlsx;*.xlsm'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='MBA file',
        initialdir="\\rsns01\maarias$\Desktop\test",
        filetypes=filetypes)

    showinfo(
        title='Selected MBA File: '+filename[-22:],
        message='Click Ok, wait and in 10 seconds the excel sheets will be created'
    )
    return filename

MBA_file_dir=MBA_file()
MBA=openpyxl.load_workbook(MBA_file_dir, data_only=True)
