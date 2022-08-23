print("Importing the information from the Ranking sheets to the final documents, please wait")
from openpyxl import Workbook
import tkinter
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
import itertools
import threading
import time
import sys
import datetime

# VARIABLES INITIATION
iZeile=2
iZeile2=2
Zeile=5
bewnr = 0

# EXTRACTING THE INFORMATION 
from extract_sheets_final import ASM
from extract_sheets_final import DDM
from extract_sheets_final import MBA

#TESTING ANIMATION
done = False
#here is the animation
def animate():
    for c in itertools.cycle(['||||||', '//////', '------', '\\\\\\\\']):
        if done:
            break
        sys.stdout.write('\rLOADING ' +' Please wait'+ c)
        sys.stdout.flush()
        time.sleep(0.1)
    sys.stdout.write('\rDone!     ')

t = threading.Thread(target=animate)
t.start()

# DEFINING THE WORKSHEETS
Ranksht=ASM["prg"]
bew=Workbook()
bews=bew.active
bews.Name='bew_to_append'
antr=Workbook()
antrs=antr.active
antrs.Name='antr_to_append'

# FUNCTION FOR CREATING THE TITLES
def titles_bew():
    bews['A1']='bewnr'
    bews['B1']='efh'
    bews['C1']='eingangdat'
    bews['D1']='eingangdat'
    bews['A1']='bewnr'
    bews['B1']='efh'
    bews['C1']='eingangdat'
    bews['D1']='fehlerkz'
    bews['E1']='mtknr'
    bews['F1']='prfzif'
    bews['G1']='bewsem'
    bews['H1']='anti'
    bews['I1']='nachname'
    bews['J1']='sortname'
    bews['K1']='vorname'
    bews['L1']='gebname'
    bews['M1']='gebort'
    bews['N1']='gebdat'
    bews['O1']='geschl'
    bews['P1']='staat'
    bews['Q1']='pozusatz'
    bews['R1']='postrasse'
    bews['S1']='poplz'
    bews['T1']='poort'
    bews['U1']='pozustbez'
    bews['V1']='pokfz'
    bews['W1']='potel'
    bews['X1']='bishsem'
    bews['Y1']='zweitst'
    bews['Z1']='hmkfzkz'
    bews['AA1']='hmkfz'
    bews['AB1']='antrnr'
    bews['AC1']='gebn'
    bews['AD1']='fehlunt'
    bews['AE1']='f1'
    bews['AF1']='f2'
    bews['AG1']='f3'
    bews['AH1']='f4'
    bews['AI1']='f5'
    bews['AJ1']='f6'
    bews['AK1']='f7'
    bews['AL1']='f8'
    bews['AM1']='bem'
    bews['AN1']='verwkz1'
    bews['AO1']='verwkz2'
    bews['AP1']='verwkz3'
    bews['AQ1']='verwkz4'
    bews['AR1']='verarbkz'
    bews['AS1']='bemlang'
    bews['AT1']='staatkez'
    bews['AU1']='anschrkz'
    bews['AV1']='kravers'
    bews['AW1']='kravnr'
    bews['AX1']='krabnr'
    bews['AY1']='berufab'
    bews['AZ1']='berufmon'
    bews['BA1']='berufjahr'
    bews['BB1']='prakt1'
    bews['BC1']='prakt2'
    bews['BD1']='sonsttaet'
    bews['BE1']='gesadauer'
    bews['BF1']='prakdauer'
    bews['BG1']='erhskfz'
    bews['BH1']='erhsart'
    bews['BI1']='erhssembrd'
    bews['BJ1']='erstsemhs'
    bews['BK1']='hssem'
    bews['BL1']='urlsem'
    bews['BM1']='praxsem'
    bews['BN1']='prax1'
    bews['BO1']='prax2'
    bews['BP1']='kolsem'
    bews['BQ1']='klinsem'
    bews['BR1']='ddrsem'
    bews['BS1']='ddrart'
    bews['BT1']='stuntsem'
    bews['BU1']='staukfz1'
    bews['BV1']='staumon1'
    bews['BW1']='stauart1'
    bews['BX1']='staukfz2'
    bews['BY1']='staumon2'
    bews['BZ1']='stauart2'
    bews['CA1']='staukfz3'
    bews['CB1']='staumon3'
    bews['CC1']='stauart3'
    bews['CD1']='wahlkz'
    bews['CE1']='wahlfb'
    bews['CF1']='antizudtxt'
    bews['CG1']='zusastrasse'
    bews['CH1']='zusaort'
    bews['CI1']='ord_kuenstname'
    bews['CJ1']='gebland'
    bews['CK1']='dokvorname'
    bews['CL1']='zvs_zusatz'
    bews['CM1']='bewnrhist'
    bews['CN1']='tnaustausch'
    bews['CO1']='titel_nachgestellt'
    bews['CP1']='basem'
    bews['CQ1']='baabdatum'
    bews['CR1']='akdsem'
    bews['CS1']='akdabdatum'
    bews['CT1']='erfassungsart'
    bews['CU1']='zustimmung_alumni'
    bews['CV1']='ersthzbart'
    bews['CW1']='ersthzbdatum'
    bews['CX1']='ersthzbkfzkz'
    bews['CY1']='ersthzbkfz'
    bews['CZ1']='ersthzbnote'
    bews['DA1']='ersthzbjahr'
    bews['DB1']='bid'
    bews['DC1']='ban'
    bews['DD1']='email'
titles_bew()

def titles_antr():
    antrs['A1']='bewnr'
    antrs['B1']='efh'
    antrs['C1']='antrnr'
    antrs['D1']='fachnr'
    antrs['E1']='kzfa'
    antrs['F1']='stg'
    antrs['G1']='abschl'
    antrs['H1']='vert'
    antrs['I1']='stuart'
    antrs['J1']='stutyp'
    antrs['K1']='stufrm'
    antrs['L1']='stgsem'
    antrs['M1']='stgspz'
    antrs['N1']='dowunsch1'
    antrs['O1']='dowunsch2'
    antrs['P1']='stort'
    antrs['Q1']='hzbwiedeu'
    antrs['R1']='hzbart'
    antrs['S1']='hzbnote'
    antrs['T1']='hzbdatum'
    antrs['U1']='hzbkfzkz'
    antrs['V1']='hzbort'
    antrs['W1']='hzbregion'
    antrs['X1']='hzbbes'
    antrs['Y1']='haerteantr'
    antrs['Z1']='haerte'
    antrs['AA1']='haertegrd'
    antrs['AB1']='haertepunkte'
    antrs['AC1']='vorzul'
    antrs['AD1']='dienst'
    antrs['AE1']='dienstende'
    antrs['AF1']='bevzul'
    antrs['AG1']='wartevor'
    antrs['AH1']='hindvor'
    antrs['AI1']='wartenach'
    antrs['AJ1']='hindnach'
    antrs['AK1']='wartemind'
    antrs['AL1']='zuskr1'
    antrs['AM1']='zusbew1'
    antrs['AN1']='zuskr2'
    antrs['AO1']='zusbew2'
    antrs['AP1']='zuskr3'
    antrs['AQ1']='zusbew3'
    antrs['AR1']='zuskr4'
    antrs['AS1']='zusbew4'
    antrs['AT1']='zuskr5'
    antrs['AU1']='zusbew5'
    antrs['AV1']='zuskr6'
    antrs['AW1']='zusbew6'
    antrs['AX1']='noteantr'
    antrs['AY1']='verbnote'
    antrs['AZ1']='notegrd'
    antrs['BA1']='noteneu'
    antrs['BB1']='zeitantr'
    antrs['BC1']='verbzeit'
    antrs['BD1']='zeitgrd'
    antrs['BE1']='wartezeit'
    antrs['BF1']='verfnote'
    antrs['BG1']='wartesem'
    antrs['BH1']='mischnote'
    antrs['BI1']='punkte'
    antrs['BJ1']='besausl'
    antrs['BK1']='messausl'
    antrs['BL1']='messzweit'
    antrs['BM1']='zulassung'
    antrs['BN1']='zuldat'
    antrs['BO1']='zulart'
    antrs['BP1']='zulfh'
    antrs['BQ1']='ablart'
    antrs['BR1']='quotenr'
    antrs['BS1']='annfrist'
    antrs['BT1']='annahme'
    antrs['BU1']='antrf1'
    antrs['BV1']='antrf2'
    antrs['BW1']='antrf3'
    antrs['BX1']='antrf4'
    antrs['BY1']='antrfu'
    antrs['BZ1']='berdatum'
    antrs['CA1']='bertaet'
    antrs['CB1']='spvo'
    antrs['CC1']='immend'
    antrs['CD1']='beweign'
    antrs['CE1']='eignnote'
    antrs['CF1']='bonussem'
    antrs['CG1']='einschreib'
    antrs['CH1']='lepsem'
    antrs['CI1']='frisem'
    antrs['CJ1']='angsemg'
    antrs['CK1']='angsems'
    antrs['CL1']='angsemb'
    antrs['CM1']='angsema'
    antrs['CN1']='klinsem'
    antrs['CO1']='kohsem'
    antrs['CP1']='eignteilnahme'
    antrs['CQ1']='eignteilnote'
    antrs['CR1']='beranfdatum'
    antrs['CS1']='klinsembean'
    antrs['CT1']='stgsembean'
    antrs['CU1']='vklinsembean'
    antrs['CV1']='bewsem'
    antrs['CW1']='kmpldatum'
    antrs['CX1']='satzid'
    antrs['CY1']='ortspraef'
    antrs['CZ1']='raliausw'
    antrs['DA1']='bewsaldo'
    antrs['DB1']='fachbind'
    antrs['DC1']='zvs_ekritpaket_01'
    antrs['DD1']='zvs_ekritpaket_02'
    antrs['DE1']='zvs_ekritpaket_03'
    antrs['DF1']='zvs_ekritpaket_04'
    antrs['DG1']='zvs_ekritpaket_05'
    antrs['DH1']='zvs_ekritpaket_06'
    antrs['DI1']='zvs_ekritpaket_07'
    antrs['DJ1']='zvs_ekritpaket_08'
    antrs['DK1']='zvs_ekritpaket_09'
    antrs['DL1']='zvs_ekritpaket_10'
    antrs['DM1']='zvs_ekritpaket_11'
    antrs['DN1']='zvs_ekritpaket_12'
    antrs['DO1']='zvs_ekritpaket_13'
    antrs['DP1']='zvs_ekritpaket_14'
    antrs['DQ1']='zvs_ekritpaket_15'
    antrs['DR1']='zvs_ekritpaket_16'
    antrs['DS1']='zvs_ekritpaket_17'
    antrs['DT1']='zvs_ekritpaket_18'
    antrs['DU1']='zvs_spezkrit'
    antrs['DV1']='zvs_ekritpaket_19'
    antrs['DW1']='zvs_ekritpaket_20'
    antrs['DX1']='los_1'
    antrs['DY1']='status_hsstart'
    antrs['DZ1']='hzb_id'
    antrs['EA1']='stg_ersatz'
    antrs['EB1']='stg_ersatz_status'
titles_antr()

# YEAR
currentDateTime = datetime.datetime.now()
date = currentDateTime.date()
year = int(date.strftime("%Y"))

# COPYING THE INFORMATION
def copying():
    global bewnr
    global bews
    global iZeile
    global iZeile2
    global Zeile
    global antrs

    while bewnr is not None:

        # 'IMMER' CELLS FOR BEWERBUNG LIST
        def immer_bew():
            bews.cell(row=iZeile,column=2).value=6731
            bews.cell(row=iZeile,column=6).value=0
            bews.cell(row=iZeile,column=25).value='N'
            bews.cell(row=iZeile,column=28).value=1
            bews.cell(row=iZeile,column=47).value='H'
            bews.cell(row=iZeile,column=62).value=0
            bews.cell(row=iZeile,column=63).value=0
            bews.cell(row=iZeile,column=64).value=0
            bews.cell(row=iZeile,column=65).value=0
            bews.cell(row=iZeile,column=68).value=0
            bews.cell(row=iZeile,column=69).value=0
            bews.cell(row=iZeile,column=70).value=0
            bews.cell(row=iZeile,column=72).value=0
            bews.cell(row=iZeile,column=91).value=0
            bews.cell(row=iZeile,column=94).value=0
            bews.cell(row=iZeile,column=96).value=0
            bews.cell(row=iZeile,column=98).value='A'
            bews.cell(row=iZeile,column=99).value='-'
            bews.cell(row=iZeile,column=104).value=0
            bews.cell(row=iZeile,column=105).value=0           
        immer_bew()

        # 'IMMER' CELLS FOR ANTR LIST
        def immer_antr():
            antrs.cell(row=iZeile2,column=2).value=6731
            antrs.cell(row=iZeile2,column=4).value=1
            antrs.cell(row=iZeile2,column=5).value='H'
            antrs.cell(row=iZeile2,column=7).value=90
            antrs.cell(row=iZeile2,column=11).value=1
            antrs.cell(row=iZeile2,column=12).value=1
            antrs.cell(row=iZeile2,column=16).value=1
            antrs.cell(row=iZeile2,column=18).value=39
            antrs.cell(row=iZeile2,column=24).value='N'
            antrs.cell(row=iZeile2,column=30).value='N'
            antrs.cell(row=iZeile2,column=92).value=0       
        immer_antr()

        # READING INFORMATION FROM RANKING TO BEW
        bewnr = Ranksht.cell(row=Zeile, column=1).value
        status= Ranksht.cell(row=Zeile, column=3).value
        eingangdat = Ranksht.cell(row=Zeile, column=6).value
        bewsem = str(year)+'2'
        nachname = Ranksht.cell(row=Zeile, column=8).value
        vorname = Ranksht.cell(row=Zeile, column=7).value
        city_study= Ranksht.cell(row=Zeile, column=45).value
        bishsem = (Ranksht.cell(row=Zeile, column=53).value)
        if bishsem is not None:
            bishsem = int(bishsem)*2
        else: 
            bishsem=8
        gebort = Ranksht.cell(row=Zeile, column=101).value
        gebdat = Ranksht.cell(row=Zeile, column=100).value
        geschl = Ranksht.cell(row=Zeile, column=13).value
        if geschl=="O":
            geschl="D" # Here the "Others" are defined as Diverse.
        staat = Ranksht.cell(row=Zeile, column=105).value
        address= Ranksht.cell(row=Zeile, column=103).value
        if address is not None:
            pozusatz = address[30:59]
            postrasse = address[0:30]
        else: 
            pozusatz = ''
            postrasse = ''
        poplz = Ranksht.cell(row=Zeile, column=107).value
        poort = Ranksht.cell(row=Zeile, column=104).value
        hmkfz = Ranksht.cell(row=Zeile, column=9).value
        ersthzbkfz = Ranksht.cell(row=Zeile, column=47).value
        if ersthzbkfz=="D":
            hmkfzkz="I"
        else: 
            hmkfzkz="A"
        erhssembrd = bewsem
        email = Ranksht.cell(row=Zeile, column=12).value

        # COPYING THIS INFORMATION INTO THE BEWERBUNGS LIST
        bews.cell(row=iZeile,column=1).value=bewnr
        bews.cell(row=iZeile,column=3).value=eingangdat
        bews.cell(row=iZeile,column=3).number_format= 'DD.MM.YYYY'
        bews.cell(row=iZeile,column=5).value=city_study
        bews.cell(row=iZeile,column=7).value=bewsem
        bews.cell(row=iZeile,column=9).value=nachname
        bews.cell(row=iZeile,column=11).value=vorname
        bews.cell(row=iZeile,column=13).value=gebort
        bews.cell(row=iZeile,column=14).value=gebdat
        bews.cell(row=iZeile,column=14).number_format= 'DD.MM.YYYY'
        bews.cell(row=iZeile,column=15).value=geschl
        bews.cell(row=iZeile,column=16).value=staat
        if status=='a'or status=='a/r':
            bews.cell(row=iZeile,column=17).value=pozusatz
            bews.cell(row=iZeile,column=18).value=postrasse
            bews.cell(row=iZeile,column=19).value=poplz
            bews.cell(row=iZeile,column=20).value=poort
            bews.cell(row=iZeile,column=108).value=email
        bews.cell(row=iZeile,column=24).value=bishsem
        bews.cell(row=iZeile,column=26).value=hmkfzkz
        if hmkfzkz=='D':
            hmkfzkz='I'
        else:
            hmkfzkz='A'
        bews.cell(row=iZeile,column=27).value=hmkfz
        bews.cell(row=iZeile,column=61).value=erhssembrd
        bews.cell(row=iZeile,column=103).value=ersthzbkfz

        # PRIORITIES
        master1 = Ranksht.cell(row=Zeile, column=118).value
        master2 = Ranksht.cell(row=Zeile, column=119).value
        master3 = Ranksht.cell(row=Zeile, column=120).value
        master4 = Ranksht.cell(row=Zeile, column=121).value
        master5 = Ranksht.cell(row=Zeile, column=122).value
        masters= [master1,master2,master3,master4,master5]
        # Changing the names for ASM.
        for i in masters:
            if i is not None:
                if i[:3]=='ASM':
                    ind=masters.index(i)
                    masters[ind]='ASM'
        # Eliminating the duplicates.
        def delete_duplicates(masters):
            new_list = []
            for n in masters:
                if not n in new_list:
                    new_list.append(n)
            return new_list
        masters = delete_duplicates(masters)
        prior=len(masters)

        # READING THE INFORMATION FROM RANKING TO ANTR
        hzbwiedeu = Ranksht.cell(row=Zeile, column=47).value
        if hzbwiedeu=='D':
            hzbwiedeu='J'
        else:
            hzbwiedeu='N'
        hzbnote = Ranksht.cell(row=Zeile, column=44).value
        hzbdatum = Ranksht.cell(row=Zeile, column=50).value
        if hzbwiedeu=='J':
            hzbkfzkz='I'
        else:
            hzbkfzkz='A' 
        hzbort = Ranksht.cell(row=Zeile, column=47).value #review, all the same?
        if status=='a'or status=='a/r':
            zulassung='J'
            zuldat = Ranksht.cell(row=Zeile, column=5).value
        else:
            zulassung='N'
            zuldat = ''
        annahme = Ranksht.cell(row=Zeile, column=3).value
        if annahme=='a':
            annahme='J'
        else:
            annahme='N'

        # COPYING THIS INFORMATION INTO THE ANTR LIST
        P=1
        while P<prior:
            immer_antr()
            antrs.cell(row=iZeile2,column=1).value=bewnr
            antrs.cell(row=iZeile2,column=3).value=P
            antrs.cell(row=iZeile2,column=6).value=masters[P-1]
            antrs.cell(row=iZeile2,column=17).value=hzbwiedeu        
            antrs.cell(row=iZeile2,column=19).value=hzbnote
            antrs.cell(row=iZeile2,column=20).value=hzbdatum
            antrs.cell(row=iZeile2,column=20).number_format= 'DD.MM.YYYY'
            antrs.cell(row=iZeile2,column=21).value=hzbkfzkz
            antrs.cell(row=iZeile2,column=22).value=hzbort
            antrs.cell(row=iZeile2,column=65).value=zulassung
            antrs.cell(row=iZeile2,column=66).value=zuldat
            antrs.cell(row=iZeile2,column=66).number_format= 'DD.MM.YYYY'
            antrs.cell(row=iZeile2,column=72).value=annahme
            antrs.cell(row=iZeile2,column=100).value=bewsem
            P=P+1
            iZeile2=iZeile2+1
        

        bewnr = Ranksht.cell(row=Zeile, column=1).value
        # Next applicant
        Zeile = Zeile+1 
        iZeile = iZeile+1


    # Deleting the last rows    
    bews.delete_rows(iZeile-1)
    antrs.delete_rows(iZeile2-1)

# FOR ASM
# Starting variables
Zeile=5
bewnr = 0
Ranksht=ASM["prg"]
copying()

# FOR DDM
# re-starting variables
Zeile=5
bewnr = 0
iZeile=iZeile-1
iZeile2=iZeile2-1
Ranksht=DDM["prg"]
copying()

# FOR MBA
# re-starting variables
Zeile=5
bewnr = 0
iZeile=iZeile-1
iZeile2=iZeile2-1
Ranksht=MBA["prg"]
copying()

# Changing the "D" for the city code. 

# Data
GS_nums =[]
for col in bews['A']:
    GS_nums.append(col.value)

country =[]
for col in bews['CY']:
    country.append(col.value)

city_study =[]
for col in bews['E']:
    city_study.append(col.value)

# Look for German Students
i = 0
GS_D = []
city_D = []
pos = [i for i, n in enumerate(country) if n == "D"]
while i <= (len(GS_nums)-1):
    if country[i] == "D":
        GS_D.append(GS_nums[i])
        city_D.append(city_study[i])
    i = i+1

window = tkinter.Tk()  # Creating the GUI
window.title("German Students County Input Tool")

# Creating title and message with number of students
title = tkinter.Label(window, text="German students County input tool ")  # Text creation
title.grid(row=0, column=0)  # Show message in GUI
num_ger_info = tkinter.Label(window, text="Number of german students found: " + str(len(GS_D)))  # Text creation
num_ger_info.grid(row=1, column=0)  # Show message in GUI


# Button function
def get_text():
    change = []
    k = 0
    l = 0
    while k <= len(TextBoxInputs)-1:
        change.append(TextBoxInputs[k].get())
        k = k+1

    for number in pos:
        country[number]= change[l]
        l = l+1
    window.destroy()
    return country




# Creating labels and Text Boxes for each german student
j = 0
TextBoxInputs = []
for x in GS_D:
    info_text = tkinter.Label(window, text="Student " + x + " lives in " + city_D[j] + " County (Landkreise) code: ")
    info_text.grid(row=j+2, column=1)

    TextBox = tkinter.Entry(window)
    TextBox.grid(row=j+2, column=2)
    TextBoxInputs.append(TextBox)
    j = j+1

# Creating the button to Save Changes
SaveButton = tkinter.Button(window, text="Save Changes", command=get_text)
SaveButton.grid(row=j+2, column=1)

window.mainloop()

i=0
while i<len(country):
    bews.cell(row=i+1,column=103).value=country[i]
    i=i+1



# Changing formats
bews.column_dimensions['C'].width = 15
bews.column_dimensions['N'].width = 15
antrs.column_dimensions['T'].width = 15
antrs.column_dimensions['BN'].width = 15

bew.save(r'\\rsns01\maarias$\Desktop\test2\bew_to_append_testing1.xlsx')
antr.save(r'\\rsns01\maarias$\Desktop\test2\antr_to_append_testing1.xlsx')

#Ending the process
done = True