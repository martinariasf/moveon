print("Importing the information from MoveOn to the Ranking sheets, please wait")
from openpyxl import Workbook, cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import openpyxl
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
import os
import itertools
import threading
import time
import sys

# VARIABLES INITIATION
Zeile= 2
iZeile=4
Cell_A = 0

# EXTRACTING THE INFORMATION 
from extract_sheets import moveon
from extract_sheets import rank
from extract_sheets import rank_file_dir
from extract_sheets import moveon_file_dir

#Making it faster for testing
#moveon=openpyxl.load_workbook(r'W:\Faculty-GS\Office-K\Ablage\08 Moveon\Hiwi - Martin Arias\Python\Excels\AcademicMoves (Wed Oct 20 2021).xlsx')
#rank=openpyxl.load_workbook(r'W:\Faculty-GS\Office-K\Ablage\08 Moveon\Hiwi - Martin Arias\Python\Excels\Ranking Template_M_v07.xlsm')




Moveonsht=moveon["Worksheet1"]
Ranksht=rank["prg"]

#FIND COUNTRY FUNCTION
Land_List = rank["Land_List"]
def funCountry(strCountry):
    for row in Land_List.iter_rows(min_row=2,min_col=3,max_col=3, max_row=250):
        for cell in row:
            if cell.value == strCountry:
                return Land_List.cell(row=cell.row,column=1).value

#FIND MASTER FUNCTION
def funMaster(strMaster):
    switcher= {
        'MBA in International Industrial Management': 'MBA',
        'MEng in Design and Development in Automotive and Mechanical Engineering (DDM)': 'DDM',
        'MEng in Automotive Systems - Car Electronics (ASM-CE)': 'ASM-CE',
        'MEng in Automotive Systems - Vehicle Dynamics (ASM-VD)': 'ASM-VD',
        'MEng in Automotive Systems - Software Based Automotive Systems (ASM-SBAS)': 'ASM-SBAS',
    }
    return switcher.get(strMaster,'')

# LOOKING FOR THE FREE CELL IN RANKING
while Ranksht.cell(row=iZeile,column=1).value is not None:
    iZeile= iZeile + 1

#TESTING ANIMATION
done = False
#here is the animation
def animate():
    for c in itertools.cycle(['||||||', '//////', '------', '\\\\\\\\']):
        if done:
            break
        sys.stdout.write('\rLOADING ' +' Please wait'+ c)
        sys.stdout.flush()
        time.sleep(0.1)
    sys.stdout.write('\rDone!     ')

t = threading.Thread(target=animate)
t.start()


# COPYING THE INFORMATION
while Cell_A is not None:

    # READING INFORMATION FROM MOVEONE
    Cell_A = Moveonsht.cell(row=Zeile, column=1).value
    Cell_B = Moveonsht.cell(row=Zeile, column=2).value
    Cell_C = Moveonsht.cell(row=Zeile, column=3).value
    Cell_D = Moveonsht.cell(row=Zeile, column=4).value
    Cell_E = Moveonsht.cell(row=Zeile, column=5).value
    Cell_F = Moveonsht.cell(row=Zeile, column=6).value
    Cell_G = Moveonsht.cell(row=Zeile, column=7).value
    Cell_H = Moveonsht.cell(row=Zeile, column=8).value
    Cell_I = Moveonsht.cell(row=Zeile, column=9).value
    Cell_J = Moveonsht.cell(row=Zeile, column=10).value
    Cell_K = Moveonsht.cell(row=Zeile, column=11).value
    Cell_L = Moveonsht.cell(row=Zeile, column=12).value
    Cell_M = Moveonsht.cell(row=Zeile, column=13).value
    Cell_N = Moveonsht.cell(row=Zeile, column=14).value
    Cell_O = Moveonsht.cell(row=Zeile, column=15).value
    Cell_P = Moveonsht.cell(row=Zeile, column=16).value
    Cell_Q = Moveonsht.cell(row=Zeile, column=17).value
    Cell_R = Moveonsht.cell(row=Zeile, column=18).value
    Cell_S = Moveonsht.cell(row=Zeile, column=19).value
    Cell_T = Moveonsht.cell(row=Zeile, column=20).value
    Cell_U = Moveonsht.cell(row=Zeile, column=21).value
    Cell_V = Moveonsht.cell(row=Zeile, column=22).value
    Cell_W = Moveonsht.cell(row=Zeile, column=23).value
    Cell_X = Moveonsht.cell(row=Zeile, column=24).value
    Cell_Y = Moveonsht.cell(row=Zeile, column=25).value
    Cell_Z = Moveonsht.cell(row=Zeile, column=26).value
    Cell_AA = Moveonsht.cell(row=Zeile, column=27).value
    Cell_AB = Moveonsht.cell(row=Zeile, column=28).value
    Cell_AC = Moveonsht.cell(row=Zeile, column=29).value
    Cell_AD = Moveonsht.cell(row=Zeile, column=30).value
    Cell_AE = Moveonsht.cell(row=Zeile, column=31).value
    Cell_AF = Moveonsht.cell(row=Zeile, column=32).value
    Cell_AG = Moveonsht.cell(row=Zeile, column=33).value
    Cell_AH = Moveonsht.cell(row=Zeile, column=34).value
    Cell_AI = Moveonsht.cell(row=Zeile, column=35).value
    Cell_AJ = Moveonsht.cell(row=Zeile, column=36).value
    Cell_AK = Moveonsht.cell(row=Zeile, column=37).value
    Cell_AL = Moveonsht.cell(row=Zeile, column=38).value
    Cell_AM = Moveonsht.cell(row=Zeile, column=39).value
    Cell_AN = Moveonsht.cell(row=Zeile, column=40).value
    Cell_AO = Moveonsht.cell(row=Zeile, column=41).value
    Cell_AP = Moveonsht.cell(row=Zeile, column=42).value
    Cell_AQ = Moveonsht.cell(row=Zeile, column=43).value
    Cell_AR = Moveonsht.cell(row=Zeile, column=44).value
    Cell_AS = Moveonsht.cell(row=Zeile, column=45).value
    Cell_AT = Moveonsht.cell(row=Zeile, column=46).value
    Cell_AU = Moveonsht.cell(row=Zeile, column=47).value
    Cell_AV = Moveonsht.cell(row=Zeile, column=48).value
    Cell_AW = Moveonsht.cell(row=Zeile, column=49).value
    Cell_AX = Moveonsht.cell(row=Zeile, column=50).value
    Cell_AY = Moveonsht.cell(row=Zeile, column=51).value
    Cell_AZ = Moveonsht.cell(row=Zeile, column=52).value
    Cell_BA = Moveonsht.cell(row=Zeile, column=53).value
    Cell_BB = Moveonsht.cell(row=Zeile, column=54).value
    Cell_BC = Moveonsht.cell(row=Zeile, column=55).value
    Cell_BD = Moveonsht.cell(row=Zeile, column=56).value
    Cell_BE = Moveonsht.cell(row=Zeile, column=57).value
    Cell_BF = Moveonsht.cell(row=Zeile, column=58).value
    Cell_BG = Moveonsht.cell(row=Zeile, column=59).value
    Cell_BH = Moveonsht.cell(row=Zeile, column=60).value
    Cell_BI = Moveonsht.cell(row=Zeile, column=61).value
    Cell_BJ = Moveonsht.cell(row=Zeile, column=62).value
    Cell_BK = Moveonsht.cell(row=Zeile, column=63).value
    Cell_BL = Moveonsht.cell(row=Zeile, column=64).value
    Cell_BM = Moveonsht.cell(row=Zeile, column=65).value
    Cell_BN = Moveonsht.cell(row=Zeile, column=66).value
    Cell_BO = Moveonsht.cell(row=Zeile, column=67).value
    Cell_BP = Moveonsht.cell(row=Zeile, column=68).value
    Cell_BQ = Moveonsht.cell(row=Zeile, column=69).value
    Cell_BR = Moveonsht.cell(row=Zeile, column=70).value
    Cell_BS = Moveonsht.cell(row=Zeile, column=71).value
    Cell_BT = Moveonsht.cell(row=Zeile, column=72).value
    Cell_BU = Moveonsht.cell(row=Zeile, column=73).value
    Cell_BV = Moveonsht.cell(row=Zeile, column=74).value
    Cell_BW = Moveonsht.cell(row=Zeile, column=75).value
    Cell_BX = Moveonsht.cell(row=Zeile, column=76).value
    Cell_BY = Moveonsht.cell(row=Zeile, column=77).value
    Cell_BZ = Moveonsht.cell(row=Zeile, column=78).value
    Cell_CA = Moveonsht.cell(row=Zeile, column=79).value
    Cell_CB = Moveonsht.cell(row=Zeile, column=80).value
    Cell_CC = Moveonsht.cell(row=Zeile, column=81).value
    Cell_CD = Moveonsht.cell(row=Zeile, column=82).value
    Cell_CE = Moveonsht.cell(row=Zeile, column=83).value
    Cell_CF = Moveonsht.cell(row=Zeile, column=84).value
    Cell_CG = Moveonsht.cell(row=Zeile, column=85).value
    Cell_CH = Moveonsht.cell(row=Zeile, column=86).value
    Cell_CI = Moveonsht.cell(row=Zeile, column=87).value
    Cell_CJ = Moveonsht.cell(row=Zeile, column=88).value
    Cell_CK = Moveonsht.cell(row=Zeile, column=89).value
    Cell_CL = Moveonsht.cell(row=Zeile, column=90).value
    Cell_CM = Moveonsht.cell(row=Zeile, column=91).value
    Cell_CN = Moveonsht.cell(row=Zeile, column=92).value
    Cell_CO = Moveonsht.cell(row=Zeile, column=93).value
    Cell_CP = Moveonsht.cell(row=Zeile, column=94).value
    Cell_CQ = Moveonsht.cell(row=Zeile, column=95).value
    Cell_CR = Moveonsht.cell(row=Zeile, column=96).value
    Cell_CS = Moveonsht.cell(row=Zeile, column=97).value
    Cell_CT = Moveonsht.cell(row=Zeile, column=98).value
    Cell_CU = Moveonsht.cell(row=Zeile, column=99).value
    Cell_CV = Moveonsht.cell(row=Zeile, column=100).value
    Cell_CW = Moveonsht.cell(row=Zeile, column=101).value
    Cell_CX = Moveonsht.cell(row=Zeile, column=102).value
    Cell_CY = Moveonsht.cell(row=Zeile, column=103).value
    Cell_CZ = Moveonsht.cell(row=Zeile, column=104).value
    Cell_DA = Moveonsht.cell(row=Zeile, column=105).value
    Cell_DB = Moveonsht.cell(row=Zeile, column=106).value
    Cell_DC = Moveonsht.cell(row=Zeile, column=107).value
    Cell_DD = Moveonsht.cell(row=Zeile, column=108).value
    Cell_DE = Moveonsht.cell(row=Zeile, column=109).value
    Cell_DF = Moveonsht.cell(row=Zeile, column=110).value
    Cell_DG = Moveonsht.cell(row=Zeile, column=111).value
    Cell_DH = Moveonsht.cell(row=Zeile, column=112).value
    Cell_DI = Moveonsht.cell(row=Zeile, column=113).value
    Cell_DJ = Moveonsht.cell(row=Zeile, column=114).value
    Cell_DK = Moveonsht.cell(row=Zeile, column=115).value
    Cell_DL = Moveonsht.cell(row=Zeile, column=116).value
    Cell_DM = Moveonsht.cell(row=Zeile, column=117).value
    Cell_DN = Moveonsht.cell(row=Zeile, column=118).value
    Cell_DO = Moveonsht.cell(row=Zeile, column=119).value
    Cell_DP = Moveonsht.cell(row=Zeile, column=120).value
    Cell_DQ = Moveonsht.cell(row=Zeile, column=121).value
    Cell_DR = Moveonsht.cell(row=Zeile, column=122).value
    Cell_DS = Moveonsht.cell(row=Zeile, column=123).value
    Cell_DT = Moveonsht.cell(row=Zeile, column=124).value
    Cell_DU = Moveonsht.cell(row=Zeile, column=125).value
    Cell_DV = Moveonsht.cell(row=Zeile, column=126).value
    Cell_DW = Moveonsht.cell(row=Zeile, column=127).value

    # COPYING THIS INFORMATION IN THE RANKING SHEETS
    Ranksht.cell(row=iZeile,column=1).value=Cell_A

    Ranksht.cell(row=iZeile,column=3).value=Cell_B

    Ranksht.cell(row=iZeile,column=7).value=Cell_D
    Ranksht.cell(row=iZeile,column=8).value=Cell_E

    Ranksht.cell(row=iZeile, column=9).value = funCountry(Cell_F)
    Ranksht.cell(row=iZeile,column=10).value = funCountry(Cell_G)

    Ranksht.cell(row=iZeile,column=12).value=Cell_H
    if Cell_I == "Männlich" :
        Ranksht.cell(row=iZeile,column=13).value = "M"
    elif Cell_I == "Weiblich":
        Ranksht.cell(row=iZeile,column=13).value = "W"
    else:
        Ranksht.cell(row=iZeile,column=13).value = "O"
   

    Ranksht.cell(row=iZeile,column=14).value=Cell_J
    Ranksht.cell(row=iZeile,column=15).value=Cell_K
    Ranksht.cell(row=iZeile,column=16).value=Cell_L

    Ranksht.cell(row=iZeile,column=17).value=Cell_M
    Ranksht.cell(row=iZeile,column=18).value=Cell_N
    Ranksht.cell(row=iZeile,column=19).value=Cell_O
    Ranksht.cell(row=iZeile,column=20).value=Cell_P

    # GMAR or GRE ?
    if Cell_Q == "GMAT (our code TJ2-P4-65)":
        Ranksht.cell(row=iZeile,column=21).value=Cell_R
        Ranksht.cell(row=iZeile,column=22).value=Cell_S
        Ranksht.cell(row=iZeile,column=23).value=Cell_T
        Ranksht.cell(row=iZeile,column=24).value=Cell_U
    else:
        Ranksht.cell(row=iZeile,column=25).value=Cell_R
        Ranksht.cell(row=iZeile,column=26).value=Cell_S
        Ranksht.cell(row=iZeile,column=27).value=Cell_T
    
    Ranksht.cell(row=iZeile,column=28).value=Cell_V
    Ranksht.cell(row=iZeile,column=29).value=Cell_W
    Ranksht.cell(row=iZeile,column=30).value=Cell_X
    Ranksht.cell(row=iZeile,column=31).value=Cell_Y
    Ranksht.cell(row=iZeile,column=32).value=Cell_Z
    Ranksht.cell(row=iZeile,column=33).value=Cell_AA

    Ranksht.cell(row=iZeile,column=45).value=Cell_AB
    Ranksht.cell(row=iZeile,column=46).value=Cell_AC

    Ranksht.cell(row=iZeile,column=47).value = funCountry(Cell_AD)

    Ranksht.cell(row=iZeile,column=48).value=Cell_AE
    Ranksht.cell(row=iZeile,column=49).value=Cell_AF
    Ranksht.cell(row=iZeile,column=50).value=Cell_AG
    Ranksht.cell(row=iZeile,column=51).value=Cell_AH
    Ranksht.cell(row=iZeile,column=52).value=Cell_AI
    Ranksht.cell(row=iZeile,column=53).value=Cell_AJ
    Ranksht.cell(row=iZeile,column=54).value=Cell_AK
    Ranksht.cell(row=iZeile,column=55).value=Cell_AL
    Ranksht.cell(row=iZeile,column=56).value=Cell_AM
    Ranksht.cell(row=iZeile,column=57).value=Cell_AN
    Ranksht.cell(row=iZeile,column=58).value=Cell_AO
    Ranksht.cell(row=iZeile,column=59).value=Cell_AP
    Ranksht.cell(row=iZeile,column=60).value=Cell_AQ
    Ranksht.cell(row=iZeile,column=61).value=Cell_AR
    Ranksht.cell(row=iZeile,column=62).value=Cell_AS
    Ranksht.cell(row=iZeile,column=63).value=Cell_AT
    Ranksht.cell(row=iZeile,column=64).value=Cell_AU
    Ranksht.cell(row=iZeile,column=65).value=Cell_AV
    Ranksht.cell(row=iZeile,column=66).value=Cell_AW
    Ranksht.cell(row=iZeile,column=67).value=Cell_AX
    Ranksht.cell(row=iZeile,column=68).value=Cell_AY
    Ranksht.cell(row=iZeile,column=69).value=Cell_AZ
    Ranksht.cell(row=iZeile,column=70).value=Cell_BA
    Ranksht.cell(row=iZeile,column=71).value=Cell_BB

    if Cell_BC == "Deutsch":
        Ranksht.cell(row=iZeile,column=72).value="German"
    else:
        Ranksht.cell(row=iZeile,column=72).value="Not German"

    Ranksht.cell(row=iZeile,column=73).value=Cell_BD
    Ranksht.cell(row=iZeile,column=74).value=Cell_BE
    Ranksht.cell(row=iZeile,column=75).value=Cell_BF
    Ranksht.cell(row=iZeile,column=76).value=Cell_BG
    Ranksht.cell(row=iZeile,column=77).value=Cell_BH
    Ranksht.cell(row=iZeile,column=78).value=Cell_BI
    Ranksht.cell(row=iZeile,column=79).value=Cell_BJ
    Ranksht.cell(row=iZeile,column=80).value=Cell_BK
    Ranksht.cell(row=iZeile,column=81).value=Cell_BL
    Ranksht.cell(row=iZeile,column=82).value=Cell_BM
    Ranksht.cell(row=iZeile,column=83).value=Cell_BN
    Ranksht.cell(row=iZeile,column=84).value=Cell_BO
    Ranksht.cell(row=iZeile,column=85).value=Cell_BP
    Ranksht.cell(row=iZeile,column=86).value=Cell_BQ
    Ranksht.cell(row=iZeile,column=87).value=Cell_BR
    Ranksht.cell(row=iZeile,column=88).value=Cell_BS
    Ranksht.cell(row=iZeile,column=89).value=Cell_BT
    Ranksht.cell(row=iZeile,column=90).value=Cell_BU
    Ranksht.cell(row=iZeile,column=91).value=Cell_BV
    Ranksht.cell(row=iZeile,column=92).value=Cell_BW
    Ranksht.cell(row=iZeile,column=93).value=Cell_BX
    Ranksht.cell(row=iZeile,column=94).value=Cell_BY
    Ranksht.cell(row=iZeile,column=95).value=Cell_BZ

    Ranksht.cell(row=iZeile,column=97).value=Cell_CA
    Ranksht.cell(row=iZeile,column=98).value=Cell_CB
    Ranksht.cell(row=iZeile,column=99).value=Cell_CC
    Ranksht.cell(row=iZeile,column=100).value=Cell_CD
    Ranksht.cell(row=iZeile,column=101).value=Cell_CE
    Ranksht.cell(row=iZeile,column=102).value=Cell_CF

    Ranksht.cell(row=iZeile,column=103).value = funCountry(Cell_CF)

    Ranksht.cell(row=iZeile,column=104).value=Cell_CG
    Ranksht.cell(row=iZeile,column=105).value=Cell_CH

    Ranksht.cell(row=iZeile,column=107).value=Cell_CI

    Ranksht.cell(row=iZeile,column=108).value = funCountry(Cell_CI)

    Ranksht.cell(row=iZeile,column=109).value=Cell_CJ
    Ranksht.cell(row=iZeile,column=110).value=Cell_CK

    Ranksht.cell(row=iZeile,column=117).value=Cell_CL
    Ranksht.cell(row=iZeile,column=118).value = funMaster(Cell_CM)
    Ranksht.cell(row=iZeile,column=119).value = funMaster(Cell_CN)
    Ranksht.cell(row=iZeile,column=120).value = funMaster(Cell_CO)
    Ranksht.cell(row=iZeile,column=121).value = funMaster(Cell_CP)
    Ranksht.cell(row=iZeile,column=122).value = funMaster(Cell_CQ)
    Ranksht.cell(row=iZeile,column=123).value=Cell_CR
    Ranksht.cell(row=iZeile,column=124).value=Cell_C


    #PAINTING
    if Ranksht.cell(row=iZeile,column=118).value == "ASM-VD":
        for rows in Ranksht.iter_cols(min_col=0, max_col=138, min_row=iZeile, max_row=iZeile):
            for cell in rows:
             cell.fill = PatternFill(start_color='d299e6',end_color='d299e6', fill_type = "solid") #LYLA - VEHICLE DYNAMIC

    if Ranksht.cell(row=iZeile,column=118).value == "ASM-CE":
        for rows in Ranksht.iter_cols(min_col=0, max_col=138, min_row=iZeile, max_row=iZeile):
            for cell in rows:
             cell.fill = PatternFill(start_color='6ab8d6',end_color='6ab8d6', fill_type = "solid") #BLUE - CAR ELECTRONICS

    if Ranksht.cell(row=iZeile,column=118).value == "ASM-SBAS":
        for rows in Ranksht.iter_cols(min_col=0, max_col=138, min_row=iZeile, max_row=iZeile):
            for cell in rows:
             cell.fill = PatternFill(start_color='94de77',end_color='94de77', fill_type = "solid") #GREEN - SOFTWARE BASED

    #PAINTING PROBLEMS
    #Problem Nr 1: Repeated participant - LIGHT RED.
    for rows in Ranksht.iter_cols(min_col=0, max_col=0, min_row=3, max_row=iZeile-2):
        for cell in rows:
            if  Ranksht.cell(row=iZeile-1,column=1).value == cell.value and Ranksht.cell(row=iZeile-1,column=1).value != "" :
               Ranksht.cell(row=iZeile-1,column=1).fill = PatternFill(start_color='fdb6b6',end_color='fdb6b6', fill_type = "solid") #LIGHT RED 

    #Problem Nr 2: Not a new participant - STRONG RED.
    if Ranksht.cell(row=iZeile-1,column=3).value != "Neue Registrierung" and Ranksht.cell(row=iZeile-1,column=3).value != "status":
        Ranksht.cell(row=iZeile-1,column=3).fill = PatternFill(start_color='fc3022',end_color='fc3022', fill_type = "solid") #STRONG RED



    Zeile=Zeile+1
    iZeile=iZeile+1

# PAINTING
#for rows in Ranksht.iter_cols(min_col=0, max_col=138, min_row=5, max_row=8):
#    for cell in rows:
#     cell.fill = PatternFill(start_color='d5947a',end_color='d5947a', fill_type = "solid") #GREEN

# SAVING
year=moveon_file_dir[-10:-6]
month=moveon_file_dir[-17:-14]
day=moveon_file_dir[-13:-11]
master_type=rank_file_dir[-8:-5]
file_name=rank_file_dir[:-20]+year+'_'+month+'_'+day+'_'+master_type+'.xlsx'
rank.save(file_name)

#MOVING THE OLD EXCEL TO THE 'alt' FOLDER
new_rank_file_dir=rank_file_dir[:-22]+'/alt/'+rank_file_dir[-22:]
os.replace(rank_file_dir, new_rank_file_dir)

#Ending the process
done = True